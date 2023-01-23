from bs4 import BeautifulSoup
import requests
import pprint
from gettext import find
from urllib import request
from bs4 import BeautifulSoup
import requests
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from  openpyxl import load_workbook

"""This file contains code for jandling the case where the websites load data dynamically (i.e with JavaScript). This file uses Selenium to get the data which solves this problem. 
Additionally, this file also contaisn the code for searching only under certain headingsin the text. We looked at the complete text, but you may use the code in this file if you want to seacrh for 
keywords under certain headings only.x"""

#create DataFrame
df = pd.DataFrame(columns=['sugar', 'sugary', 'C', 'D', 'E'])

# Place the url of the website that you want to search for here.
url = 'https://inquestmed.com/articles/why-type-2-diabetes-is-dangerous/'


def find_all(a_str, sub):
    """Searches for a substring in given text, and returns the indices of all instances found in the text (if any)"""
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub) # use start += 1 to find overlapping matches

def split_into_sentences(text):
    alphabets= "([A-Za-z])"
    prefixes = "(Mr|St|Mrs|Ms|Dr)[.]"
    suffixes = "(Inc|Ltd|Jr|Sr|Co)"
    starters = "(Mr|Mrs|Ms|Dr|He\s|She\s|It\s|They\s|Their\s|Our\s|We\s|But\s|However\s|That\s|This\s|Wherever)"
    acronyms = "([A-Z][.][A-Z][.](?:[A-Z][.])?)"
    websites = "[.](com|net|org|io|gov)"
    digits = "([0-9])"

    def split(text):
        text = " " + text + "  "
        text = text.replace("\n"," ")
        text = re.sub(prefixes,"\\1<prd>",text)
        text = re.sub(websites,"<prd>\\1",text)
        text = re.sub(digits + "[.]" + digits,"\\1<prd>\\2",text)
        if "..." in text: text = text.replace("...","<prd><prd><prd>")
        if "Ph.D" in text: text = text.replace("Ph.D.","Ph<prd>D<prd>")
        text = re.sub("\s" + alphabets + "[.] "," \\1<prd> ",text)
        text = re.sub(acronyms+" "+starters,"\\1<stop> \\2",text)
        text = re.sub(alphabets + "[.]" + alphabets + "[.]" + alphabets + "[.]","\\1<prd>\\2<prd>\\3<prd>",text)
        text = re.sub(alphabets + "[.]" + alphabets + "[.]","\\1<prd>\\2<prd>",text)
        text = re.sub(" "+suffixes+"[.] "+starters," \\1<stop> \\2",text)
        text = re.sub(" "+suffixes+"[.]"," \\1<prd>",text)
        text = re.sub(" " + alphabets + "[.]"," \\1<prd>",text)
        if "”" in text: text = text.replace(".”","”.")
        if "\"" in text: text = text.replace(".\"","\".")
        if "!" in text: text = text.replace("!\"","\"!")
        if "?" in text: text = text.replace("?\"","\"?")
        text = text.replace(".",".<stop>")
        text = text.replace("?","?<stop>")
        text = text.replace("!","!<stop>")
        text = text.replace("<prd>",".")
        sentences = text.split("<stop>")
        sentences = sentences[:-1]
        sentences = [s.strip() for s in sentences]
        return sentences


def exclude_matches(text, term, positions):
    new_positions = []
    for position in positions:
        len_term = len(term)
        start_position = position - len_term -1
        end_position = start_position + len_term

        if (text[start_position:end_position]) != "blood":
            new_positions.append(position)
    
    return new_positions

# for finding all h tags in soup
heading_tags = ["h1", "h2"]
def get_tags(tags_list):
    for tags in soup.find_all(tags_list):
        print(tags.name + ' -> ' + tags.text.strip())


upper_tagrget_headings = []

# return the html for the given url
def get_html(url):
    opt = Options()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opt)
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    return soup

# for finding all the tags in html
def get_tags(soup):
    headings = soup.find_all(re.compile("^h[1-6]$"))
    return headings

html = get_html(url)

headings=[]

for heading in get_tags(html):
    headings.append(heading)

def contains(list, string):
    for item in list:
        if string.find(item) != -1:
            return True

# list of terms to search
search_terms = ['prevent', 'Prevent','cause', 'Cause', 'control', 'Control', 'risk', 'Risk', 'risks','risk factor', 'Risk', 'nutrition', 'causes', 'Causes']

target_headings = []
for item in headings:
    if contains(search_terms, item.text):
        target_headings.append(item)

print(target_headings)

final_data = []
for target_heading in target_headings:
    siblings = target_heading.find_next_siblings()
    
    text = ''
    for sib in siblings:
        if sib.name in heading_tags:
            break
        else:
            text += sib.text
    
    final_data.append({'heading':target_heading, 'text': text})


the_text = ''
for x in final_data:
    the_text += x['text']




search_items = ['sugar','sugary', 'sweets', 'sweetened', 'sweetener','soda','sucrose','fructose', 'glycemic index', 'glycemic load', 'processed', 'refined', 'fiber', 'whole grain', 'whole-grain', 
'fruit', 'vegetable', 'saturated fat', 'unsaturated fat', 'low-fat', 'low fat', 'red meat', 'processed meat', 'obese', 'obesity', 'weight', 'abdominal fat', 'belly fat', 'body mass index', 'bmi', 
'adiposity', 'exercis', 'activ', 'sedentary', 'hypertension', 'blood pressure', 'blood pressure', 'dyslipidemia',  'triglyceride', 'cholesterol', 'ldl', 'hdl', 'age', 'old', 'history', 'genetic', 'ethnicit']


def read_column(filepath, cell, rows):
    workbook = load_workbook(filepath)

    worksheet = workbook.active

    list_of_websites = []
    for row in range(1, rows):
        column = cell
        cell_number = column + str(row)
        cell_value = worksheet[cell_number].value
        list_of_websites.append(cell_value)

    list_of_websites.pop(0)
    return(list_of_websites)


def do_all(url):
    
    upper_tagrget_headings = []

    html = get_html(url)

    headings=[]

    for heading in get_tags(html):
        headings.append(heading)

    # list of terms to search
    search_terms = ['prevent', 'Prevent','cause', 'Cause', 'control', 'Control', 'risk', 'Risk', 'risks','risk factor', 'Risk', 'nutrition', 'causes', 'Causes']

    target_headings = []
    for item in headings:
        if contains(search_terms, item.text):
            target_headings.append(item)

    print(target_headings)

    final_data = []
    for target_heading in target_headings:
        siblings = target_heading.find_next_siblings()
        
        text = ''
        for sib in siblings:
            if sib.name in heading_tags:
                break
            else:
                text += sib.text
        
        final_data.append({'heading':target_heading, 'text': text})


    the_text = ''
    for x in final_data:
        the_text += x['text']

    columns = []
    for search_term in search_items:
        columns.append(search_term)

    #create DataFrame
    df = pd.DataFrame(columns=columns)

    lowercase_text = the_text.lower()
    print(lowercase_text)

    word_counts = {}
    for item in search_items:

        matches = list(find_all(lowercase_text, item))

        refined_matches = exclude_matches(lowercase_text, 'blood', matches)

        try:
            if len(refined_matches) > 0:
                word_counts[item] = 1
        except KeyError:
            if len(refined_matches) > 0:
                word_counts[item] = 1

    print(word_counts)

    df_row = {}
    df_row.update({'Website': url})
    for (column_name, column_data) in df.iteritems():
        if column_name in word_counts.keys():
            df_row[column_name] = 1
        else:
            df_row[column_name] = 0
        
        if len(target_headings) == 0:
            df_row[column_name] = 'No Hs'
        
        if len(word_counts) == 0 and len(target_headings) != 0:
            df_row[column_name] = 'No Ws'

    df_row_list = []
    for key, value in df_row.items():
        df_row_list.append(value)

    wb = openpyxl.load_workbook("output.xlsx") 
    
    sheet = wb.active 

    sheet.append(df_row_list)
    
    wb.save('output.xlsx')

    print('DataFrame is written successfully to Excel File.')


